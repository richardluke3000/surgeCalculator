from openpyxl import load_workbook

workbook = load_workbook( filename= "D:\projects\extract\Rumphi Surge Call Data 31 May 2021\Rumphi Surge Call Data 31 May 2021.xlsx" )

print (workbook.sheetnames)

# sheet = workbook.active

# print( sheet.title )

# print ( sheet ["a1"].value )

# if sheet ["a1"].value == "OPD Screening -- Biweekly Reporting" :
#     # for value in sheet[5:6]:
#     #     print(value.value)
#     for value in sheet.iter_rows(min_row= 6,
#                                  max_row=12,
#                                  min_col=1,
#                                  max_col=5,
#                                  values_only=True):
#         print(value)
    
