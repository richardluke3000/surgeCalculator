from openpyxl import load_workbook
import datetime

datetime.date(2020,4,2)

# for getting cmd arguments
import sys

from openpyxl.workbook.workbook import Workbook

# collumn and row configurations

min_row= 6,
max_row=100,
max_col=100,
min_col=1,
# end row and collumn configuration

# file location
file = sys.argv[1]

# new file savename
filename = sys.argv[2]

# sheet index to extract from
sheet_index = int(sys.argv[3])

# load the file from sytem
workbook = load_workbook( filename= file )

# print (workbook.sheetnames)
sheet = workbook._active_sheet_index = sheet_index;


# print(sheet)
# print( workbook.active.title )
# print(workbook.active)
# sheet
# print (sheet)





# print( sheet.title )

# print ( sheet ["a1"].value )

# compare if the selected index is actually the right sheet
if True :
    
    
    # values for starting point of new excel document
    x = 1
    y = 1
    
    # new workbook to extract in
    workbook2 = Workbook()
    
    # actuve sheet
    sheet2 = workbook.active

    # iterate over the old excel file with custom starting/ending rows and collumns
    for value in sheet2.iter_rows(min_row= 6,
                                 max_row=100,
                                 max_col=100,
                                 min_col=1,
                                 values_only=True):
        # extract every single value in a cell
        
        for cell in value:
            box = workbook2.active.cell(row=x, column=y, value=cell)
            y += 1
        x+=1
        y = 1
        
        # for x in value:
        #     sheet = x
        # sheet['A1'] = value[0]
        
    

    workbook2.save(filename)
    print('succesful')
    
    
