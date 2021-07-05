from openpyxl import load_workbook
import datetime

from openpyxl.workbook.workbook import Workbook

# for getting cmd arguments
import sys


# python dates.py "d:\projects\extract\Rumphi Surge Call Data 31 May 2021\Rumphi Surge Call Data 31 May 2021.xlsx" date.xlsx 5

# column and row configurations

min_row= int( sys.argv[5] )
max_row= int( sys.argv[6] )
max_col= int( sys.argv[7] )
min_col= int( sys.argv[8] )
# end row and collumn configuration


# file location
file = sys.argv[1]

# new file savename
filename = sys.argv[2]

# sheet index to extract from
sheet_index = int(sys.argv[3])

datesplit = sys.argv[4].split('-')
year = int(datesplit[0])
month = int( datesplit[1] )
date = int ( datesplit[2] )
customstartdate  = datetime.datetime(year, month, date)

# load the file from sytem
workbook = load_workbook( filename= file )

# print (workbook.sheetnames)
sheet = workbook._active_sheet_index = sheet_index;




# print ( sheet ["a1"].value )

# compare if the selected index is actually the right sheet
if True :
    
    
    # values for starting point of new excel document
    x = 1
    y = 1
    
    # new workbook to extract in
    workbook2 = Workbook()
    
    # actuve sheet
    sheet2 = workbook.active

    # iterate over the old excel file with custom starting/ending rows and collumns
    for value in sheet2.iter_rows(min_row= (min_row - 2),
                                 max_row=max_row,
                                 max_col=max_col,
                                 min_col=min_col,
                                 values_only=True):
        # EXTRACT HEADERS
        if ( x <= 2) :
            for cell in value:
                workbook2.active.cell(row=x, column=y, value=cell)
                y += 1
            x += 1
            y = 1    
            continue
           
        
        
        # extract every single value in a cell if the date value is true for the chosen filter
        if( (type( sheet2['E'+ str(min_row) ].value  ) ==  type(customstartdate)) and  (sheet2['E'+ str(min_row) ].value == customstartdate ) ):
            
            
            for cell in value:
                workbook2.active.cell(row=x, column=y, value=cell)
                y += 1
            x +=1
            y = 1
            min_row += 1
        else :
            # x +=1
            y = 1
            min_row += 1
       
        
    

    workbook2.save(filename)
    
    print('successful', file=sys.stdout)
    
    
